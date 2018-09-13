import sys
sys.path.append("E:/test/dcms/ChengGuan")
import time
import requests
import openpyxl
from common.constant_all import getConstant
from common.writeAndReadText import writeAndReadTextFile





def readData():
    wb = openpyxl.load_workbook('E:/test/dcms/ChengGuan/testFile/newPersonnel/newPresonnel_2.xlsx')
    # ws = wb['Sheet1']
    ws = wb.active

    
    i = 2
    personData = {}
    personData['max_row'] = ws.max_row
    personData[ws["A" + str(i-1)].value] = ws["A" + str(i)].value
    personData[ws["B" + str(i-1)].value] = ws["B" + str(i)].value
    personData[ws["C" + str(i-1)].value] = ws["C" + str(i)].value
    personData[ws["D" + str(i-1)].value] = ws["D" + str(i)].value
    personData[ws["E" + str(i-1)].value] = ws["E" + str(i)].value
    personData[ws["F" + str(i-1)].value] = ws["F" + str(i)].value
    personData[ws["G" + str(i-1)].value] = ws["G" + str(i)].value
    personData[ws["H" + str(i-1)].value] = ws["H" + str(i)].value
    personData[ws["I" + str(i-1)].value] = ws["I" + str(i)].value
    personData[ws["J" + str(i-1)].value] = ws["J" + str(i)].value
    personData[ws["K" + str(i-1)].value] = ws["K" + str(i)].value
    personData[ws["L" + str(i-1)].value] = ws["L" + str(i)].value

    return personData

def newPersonnel(obj):
    # print(obj['mobilephone'])
    url = getConstant.IP_WEB_180+'/dcms/bmsAdmin/Admin-save.action'
    data = {
        "tempJumpUrl":"",
        "id":"",	
        "groups":"",
        "logonpassword":"",	
        "deptId":obj['deptId'],
        "bgids":"",	
        "bgnms":"",	
        "name":	obj['name'],
        "logonname":obj['logonname'],
        "password":obj['password'],
        "position":obj['position'],
        "phone":obj['phone'],
        "mobilephone":obj['mobilephone'],
        "mac":"",
        "num":obj['num'],
        "deptname":obj['deptname'],
        "jobposition":obj['jobposition'],
        "lead":"",	
        "activation":obj['activation'],
        "zxcode":"",	
        "zxphone":"",	
        "eno":obj['eno'],
        "file":	"",
        "bgName":"",	
        "groupnames":"",	
    }
    header = {
        "Cookie":writeAndReadTextFile().test_readCookies()
    }
    r = requests.post(url = url,data = data,headers = header,allow_redirects=False)
    # print(r.text)
    if ('Set-Cookie' not in r.headers or '用户名已经存在' in r.text):
        print("返回结果",r.text)
        filename = 'E:/test/dcms/ChengGuan/testFile/newPersonnel/newPresonnel_2.xlsx'
        wb = wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws.delete_rows(2,1) #删除index为2后面的2行
        wb.save(filename)
        print("删除成功")

if __name__ == "__main__":
    
    personData = readData()
    for i in range(1,personData['max_row']): #personData['max_row']
        time.sleep(1)
        newPersonnel(readData())



















