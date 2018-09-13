import sys
sys.path.append("E:/test/dcms/ChengGuan")
import time
import requests
import openpyxl
from common.constant_all import getConstant
from common.writeAndReadText import writeAndReadTextFile



class readExcel():

    def __init__(self,filename):

        self.filename = filename

        # self.i = i

    def readData(self,i):
        self.i = i
        wb = openpyxl.load_workbook(filename)
        # ws = wb['Sheet1']
        ws = wb.active

        
        j = 1
        personData = {}
        personData['max_row'] = ws.max_row
        personData[ws["A" + str(j)].value] = ws["A" + str(self.i)].value
        personData[ws["B" + str(j)].value] = ws["B" + str(self.i)].value
        personData[ws["C" + str(j)].value] = ws["C" + str(self.i)].value
        personData[ws["D" + str(j)].value] = ws["D" + str(self.i)].value
        personData[ws["E" + str(j)].value] = ws["E" + str(self.i)].value
        personData[ws["F" + str(j)].value] = ws["F" + str(self.i)].value
        personData[ws["G" + str(j)].value] = ws["G" + str(self.i)].value
        personData[ws["H" + str(j)].value] = ws["H" + str(self.i)].value
        personData[ws["I" + str(j)].value] = ws["I" + str(self.i)].value
        personData[ws["J" + str(j)].value] = ws["J" + str(self.i)].value
        personData[ws["K" + str(j)].value] = ws["K" + str(self.i)].value
        personData[ws["L" + str(j)].value] = ws["L" + str(self.i)].value
        print(personData)

        return personData

    def newPersonnel(self,obj):
        # print("obj",obj)
        url = getConstant.IP_WEB_180+'/dcms/bmsAdmin/Admin-save.action'
        data = {
            "tempJumpUrl":"",
            "id":"",	
            "groups":"",
            "logonpassword":"",	
            "deptId":obj['deptId'],
            "bgids":"",	
            "bgnms":"",	
            "name":	obj['name'],
            "logonname":obj['logonname'],
            "password":obj['password'],
            "position":obj['position'],
            "phone":obj['phone'],
            "mobilephone":obj['mobilephone'],
            "mac":"",
            "num":obj['num'],
            "deptname":obj['deptname'],
            "jobposition":obj['jobposition'],
            "lead":"",	
            "activation":obj['activation'],
            "zxcode":"",	
            "zxphone":"",	
            "eno":obj['eno'],
            "file":	"",
            "bgName":"",	
            "groupnames":"",	
        }
        header = {
            "Cookie":writeAndReadTextFile().test_readCookies()
        }
        r = requests.post(url = url,data = data,headers = header,allow_redirects=False)
        # print("返回值：",r.text)
        if ('Set-Cookie' in r.headers):
            print("请您先登录")
        elif '用户名已经存在' in r.text:
            print("返回结果",r.text)
        else:
            print("返回结果",r.text)

if __name__ == "__main__":
    filename = 'E:/test/dcms/ChengGuan/testFile/newPersonnel/newPresonnel_2.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # rowNumber = ws.max_row
    Excel = readExcel(filename)
    # personData = readData()
    for i in range(2,ws.max_row):
        time.sleep(1)
        Excel.newPersonnel(Excel.readData(i))



















