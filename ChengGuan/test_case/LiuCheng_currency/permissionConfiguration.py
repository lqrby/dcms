import sys
sys.path.append("E:/test/dcms/ChengGuan")
import time
import requests
import openpyxl
from common.constant_all import getConstant
from common.writeAndReadText import writeAndReadTextFile


#用户权限配置
class PermissionConfiguration():

    def __init__(self,filename):

        self.filename = filename
        self.header = {
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36",
            "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
            "Accept-Encoding":"gzip, deflate",
            "Accept-Language":"zh-CN,zh;q=0.9,en;q=0.8",
            "Cookie":writeAndReadTextFile().test_readCookies()
            }

        if '180' in getConstant.IP:
            self.ip = getConstant.IP+getConstant.PORT_7897
        else:
            self.ip = getConstant.IP

    #权限设置
    def authorityAllocation(self,i):
        qxszurl = self.ip+"/bmsAdmin/UserRole!saveUserRole.action"
        qxszdata = {
            "userId":"8a8a8483693418b0016936f2a4501d2a",
            "roleIds":"402883845f295831015f296be837003d,8a8a848361464faa01615438e5051c37,8a8a848260e4cad20160e80d173a1a61,8a8a848261464b200161543976571a61,8a8a8482645fd43301645fd8461c0020,402883835baa13f2015bad850b66000c,8a8a84835f9608f8015fa4f582f61c5d,8a8a84835adcc7b3015ba3ec8296345f",
            "tempJumpUrl":"Admin-getDeptUser.action?deptId=undefined",
            "checkbox":"402883845f295831015f296be837003d",       #巡检抽查
            "checkbox":"8a8a848361464faa01615438e5051c37",       #论坛
            "checkbox":"8a8a848260e4cad20160e80d173a1a61",       #油烟污染
            "checkbox":"8a8a848261464b200161543976571a61",       #督查考核
            "checkbox":"8a8a8482645fd43301645fd8461c0020",       #静态交通
            "checkbox":"402883835baa13f2015bad850b66000c",       #督办指导
            "checkbox":"8a8a84835f9608f8015fa4f582f61c5d",       #违法建筑
            "checkbox":"8a8a84835adcc7b3015ba3ec8296345f"        #领导
        }
        r = requests.post(url = url,data = data,headers = self.header,allow_redirects=False)
        # print("返回值：",r.text)
        if ('Set-Cookie' in r.headers):
            print("请您先登录")
        elif '用户名已经存在' in r.text:
            print("返回结果",r.text)
        else:
            print("返回结果",r.text)

if __name__ == "__main__":
    filename = 'E:/test/dcms/ChengGuan/testFile/newPersonnel/newPresonnel.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # rowNumber = ws.max_row
    Excel = readExcel(filename)
    # personData = readData()
    for i in range(2,ws.max_row+1):
        time.sleep(1)
        Excel.newPersonnel(Excel.readData(i))



















