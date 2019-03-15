import sys
sys.path.append("E:/test/dcms/ChengGuan")
import time
import requests
import openpyxl
from common.constant_all import getConstant
from common.writeAndReadText import writeAndReadTextFile



class readExcel():

    def __init__(self,filename):

        self.filename = filename
        self.header = {
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36",
            "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
            "Accept-Encoding":"gzip, deflate",
            "Accept-Language":"zh-CN,zh;q=0.9,en;q=0.8",
            "Cookie":writeAndReadTextFile().test_readCookies()
            }
        if '180' in getConstant.IP:
            self.ip = getConstant.IP+getConstant.PORT_7897
        else:
            self.ip = getConstant.IP

    def readData(self,i):
        self.i = i
        wb = openpyxl.load_workbook(filename)
        # ws = wb['Sheet1']
        ws = wb.active

        
        j = 1
        personData = {}
        personData['max_row'] = ws.max_row
        personData[ws["A" + str(j)].value] = ws["A" + str(self.i)].value
        personData[ws["B" + str(j)].value] = ws["B" + str(self.i)].value
        personData[ws["C" + str(j)].value] = ws["C" + str(self.i)].value
        personData[ws["D" + str(j)].value] = ws["D" + str(self.i)].value
        personData[ws["E" + str(j)].value] = ws["E" + str(self.i)].value
        personData[ws["F" + str(j)].value] = ws["F" + str(self.i)].value
        personData[ws["G" + str(j)].value] = ws["G" + str(self.i)].value
        personData[ws["H" + str(j)].value] = ws["H" + str(self.i)].value
        personData[ws["I" + str(j)].value] = ws["I" + str(self.i)].value
        personData[ws["J" + str(j)].value] = ws["J" + str(self.i)].value
        personData[ws["K" + str(j)].value] = ws["K" + str(self.i)].value
        personData[ws["L" + str(j)].value] = ws["L" + str(self.i)].value
        print(personData)

        return personData

    def newPersonnel(self,obj):
        url = self.ip+'/dcms/bmsAdmin/Admin-save.action'
        data = {
            "tempJumpUrl":"",
            "id":"",	
            "groups":"",
            "logonpassword":"",	
            "deptId":obj['deptId'],
            "bgids":"",	
            "bgnms":"",	
            "name":	obj['name'],
            "logonname":obj['logonname'],
            "password":obj['password'],
            "position":obj['position'],
            "phone":obj['phone'],
            "mobilephone":obj['mobilephone'],
            "mac":"",
            "num":obj['num'],
            "deptname":obj['deptname'],
            "jobposition":obj['jobposition'],
            "lead":"",	
            "activation":obj['activation'],
            "zxcode":"",	
            "zxphone":"",	
            "eno":obj['eno'],
            "file":	"",
            "bgName":"",	
            "groupnames":"",	
        }
        r = requests.post(url = url,data = data,headers = self.header,allow_redirects=False,timeout = 20)
        # print("返回值：",r.text)
        if ('Set-Cookie' in r.headers):
            print("请您先登录")
        elif '用户名已经存在' in r.text:
            print("返回结果",r.text)
        else:
            print("返回结果",r.text)
        r.connection.close()

if __name__ == "__main__":
    filename = 'E:/test/dcms/ChengGuan/testFile/newPersonnel/newPresonnel.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # rowNumber = ws.max_row
    Excel = readExcel(filename)
    # personData = readData()
    for i in range(2,ws.max_row+1):
        time.sleep(1)
        Excel.newPersonnel(Excel.readData(i))



















