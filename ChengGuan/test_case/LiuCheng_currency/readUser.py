import unittest
import sys
sys.path.append("E:/test/dcms/ChengGuan/")
import time,random,json,requests
import threading
from config.Log import logging
from selenium import webdriver
from login import allLogin
from jobEntry import submitOrder
from queRen import confirm
from heShi import verify
from liAn import setUpCase
from paiFa import distribution
from guaZhang import hangUp
from chuLi import fileFandling
from fuHeAndHuiFang import reviewAndReturnVisit
from common.writeAndReadText import writeAndReadTextFile
from common.constant_all import getConstant
from new_personnel_zx180 import readExcel
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
from common.writeAndReadText import writeAndReadTextFile
from chengguan_authCode import test_login_authCode
from common.constant_all import getConstant

class read_user():

    def read_ZFJ_User(self):
        if '180' in getConstant.IP:
            # ip = getConstant.IP+getConstant.PORT_7897
            filename = 'E:/test/dcms/ChengGuan/testFile/newPersonnel/zfj_user_180.xlsx'
        else:
            # ip = getConstant.IP
            filename = 'E:/test/dcms/ChengGuan/testFile/newPersonnel/zfj_user_180.xlsx'
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        return ws.max_row
        # for i in range(2,ws.max_row+1):
        #     personData = readExcel(filename).readData(i)
        #     url= ip+"/dcms/PwasAdmin/mobile-loginadmin.action"
        #     print("url:",url)
        #     zfj_data = {
        #         "role":personData['jobposition'],
        #         "logonname":personData['logonname'],
        #         "logonpassword":personData['password']
        #     }  
        #     res = requests.post(url,zfj_data).text
        #     zfjResult = json.loads(res)
        #     if 'message' in zfjResult and zfjResult['message'] == 'success':
        #         print("执法局apk:登录成功")
        #         # return zfjResult
        #     else:
        #         print("执法局apk:登录失败！！！",zfjResult)
        #         # #return False
if __name__ == "__main__":

    read_user().read_ZFJ_User()